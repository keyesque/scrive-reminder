from openpyxl import load_workbook
import json

# Constants
target_color = "FFFF0000"
document_name = "Arbetsdokument.xlsx"
document_sheet = "Sheet1"

# Variables
wb = load_workbook(document_name)
ws = wb[document_sheet]
recipients = []

# Iterate through excel
for row in ws.iter_rows(min_row=2):
    name = row[5].value
    # Use first name
    name = name.split(" ")[0];

    email = row[4].value
    id = row[3]
    cell = row[4]
    color = cell.fill.start_color.rgb

    if color == target_color:
        recipient = {
            "Name": name,
            "Email": email,
            "Id": id
        }
        recipients.append(recipient)

# Format to JSON
json_file = json.dumps(recipients)
print(json_file)

# Write to file
with open("recipients.json", "w") as f:
    f.write(json_file)
